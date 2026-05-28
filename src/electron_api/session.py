from __future__ import annotations

import shutil
import os
from pathlib import Path
from typing import Any

import globals
import packer
from config_manager import CONFIG_FILE, ConfigManager
from globals import COLOR_MAP, LANGUAGE_MAP, UNIQUENESS_IDS, WORKING_DIR
from inventory_handler import InventoryHandler
from relic_checker import InvalidReason, RelicChecker, is_curse_invalid
from source_data_handler import SourceDataHandler
from utils.backup import create_backup
from vessel_handler import LoadoutHandler, is_vessel_available


EMPTY_EFFECT_IDS = {0, -1, 0xFFFFFFFF}


class SaveEditorSession:
    """UI-free session facade for Electron.

    Existing save-editing modules still use process-global state. This class
    makes that state explicit at the Electron boundary without importing
    Tkinter-heavy `Final.py`.
    """

    def __init__(self, work_dir: Path | None = None):
        base_work_dir = Path(os.environ.get("NIGHTREIGN_ELECTRON_WORK_DIR", WORKING_DIR))
        self.work_dir = Path(work_dir or base_work_dir / "decrypted_output_electron")
        self.import_work_dir = Path(base_work_dir / "decrypted_output_electron_import")
        self.backup_dir = Path(base_work_dir / "backup")
        self.config = ConfigManager()
        self.game_data = SourceDataHandler(language=self.config.language)
        self.relic_checker = RelicChecker()
        self.inventory = InventoryHandler()
        self.loadouts = LoadoutHandler()
        self.save_file_path: Path | None = None
        self.userdata_path: Path | None = None
        self.characters: list[tuple[str, Path]] = []
        self.import_save_file_path: Path | None = None
        self.import_characters: list[tuple[str, Path]] = []
        self.pending_loadout_import: Any | None = None
        self.pending_loadout_import_snapshot: bytearray | None = None
        self.selected_character_index: int | None = None

    def ping(self) -> dict[str, Any]:
        return {"ok": True, "language": self.config.language}

    def get_settings(self) -> dict[str, Any]:
        return self._serialize_settings()

    def update_settings(self, settings: dict[str, Any]) -> dict[str, Any]:
        selected_character = None
        language_changed = False

        if "language" in settings:
            language = str(settings["language"])
            if language not in LANGUAGE_MAP:
                raise ValueError(f"Unsupported language: {language}")
            if language != self.config.language:
                if not self.game_data.reload_text(language):
                    raise ValueError(f"Language resources could not be loaded: {language}")
                self.config.language = language
                language_changed = True

        if "theme" in settings:
            theme = str(settings["theme"])
            if theme not in {"Light", "Dark"}:
                raise ValueError("Theme must be Light or Dark")
            self.config.theme = theme

        if "reduceMessagePop" in settings:
            self.config.reduce_message_pop = bool(settings["reduceMessagePop"])

        if "autoBackup" in settings:
            self.config.auto_backup = bool(settings["autoBackup"])

        if "maxBackups" in settings:
            max_backups = int(settings["maxBackups"])
            if max_backups < 0 or max_backups > 100:
                raise ValueError("Max backups must be between 0 and 100")
            self.config.max_backups = max_backups

        if language_changed and globals.data is not None and self.userdata_path is not None:
            selected_character = self._current_character_state()

        return {
            "settings": self._serialize_settings(),
            "selectedCharacter": selected_character,
        }

    def open_save(self, file_path: str, preferred_character_index: int | None = None) -> dict[str, Any]:
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"Save file not found: {path}")

        self._reset_work_dir(self.work_dir)
        packer.unpack(path, self.work_dir)
        self.save_file_path = path
        self.config.last_file = str(path)
        self.characters = self._name_to_path(self.work_dir)

        selected = None
        if self.characters:
            selected_index = 0
            if preferred_character_index is not None and 0 <= preferred_character_index < len(self.characters):
                selected_index = preferred_character_index
            selected = self.load_character(selected_index)

        return {
            "savePath": str(path),
            "mode": packer.detect_repacker(self.work_dir).mode,
            "unpackDir": str(self.work_dir),
            "characters": self._serialize_characters(),
            "selectedCharacter": selected,
        }

    def open_last_save(self) -> dict[str, Any] | None:
        last_file = str(self.config.last_file or "")
        if not last_file:
            return None

        path = Path(last_file)
        if not path.is_file():
            return None

        return self.open_save(str(path), preferred_character_index=int(self.config.last_char_index))

    def load_character(self, index: int) -> dict[str, Any]:
        if not self.characters:
            raise RuntimeError("No save file is open")
        if index < 0 or index >= len(self.characters):
            raise IndexError(f"Invalid character index: {index}")

        name, path = self.characters[index]
        self.userdata_path = path
        self.selected_character_index = index
        self.config.last_char_index = index
        globals.data = bytearray(path.read_bytes())

        self.inventory.parse()
        self.loadouts.parse()
        self.inventory.set_illegal_relics()

        return self._current_character_state()

    def open_import_save(self, file_path: str) -> dict[str, Any]:
        self._require_open_save()
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"Save file not found: {path}")

        self._reset_work_dir(self.import_work_dir)
        packer.unpack(path, self.import_work_dir)
        self.import_save_file_path = path
        self.import_characters = self._name_to_path(self.import_work_dir)

        return {
            "savePath": str(path),
            "mode": packer.detect_repacker(self.import_work_dir).mode,
            "unpackDir": str(self.import_work_dir),
            "characters": self._serialize_import_characters(),
        }

    def replace_character(self, import_index: int) -> dict[str, Any]:
        self._require_character()
        if not self.import_characters:
            raise RuntimeError("No import save file is open")
        if import_index < 0 or import_index >= len(self.import_characters):
            raise IndexError(f"Invalid import character index: {import_index}")
        assert globals.data is not None

        _name, import_path = self.import_characters[import_index]
        imported_data = import_path.read_bytes()
        current_size = len(globals.data)
        if len(imported_data) <= current_size:
            globals.data = bytearray(imported_data + bytes(globals.data[len(imported_data):]))
        else:
            globals.data = bytearray(imported_data[:current_size])

        self.save_current_character()
        self._reparse_current_character()
        self._refresh_characters_preserving_selection()
        return self._current_character_state()

    def get_stats(self) -> dict[str, int]:
        self._require_character()
        return {
            "murks": self.inventory.murks,
            "sigs": self.inventory.sigs,
        }

    def update_stat(self, field: str, value: int) -> dict[str, Any]:
        self._require_character()
        if field not in {"murks", "sigs"}:
            raise ValueError(f"Unknown stat field: {field}")
        self._validate_uint32(value, "Stat value")

        if field == "murks":
            self.inventory.murks = value
        else:
            self.inventory.sigs = value

        self.save_current_character()
        self._reparse_current_character()
        return self._current_character_state()

    def save_current_character(self) -> dict[str, Any]:
        self._require_character()
        assert globals.data is not None
        assert self.userdata_path is not None
        self.userdata_path.write_bytes(globals.data)
        return {
            "path": str(self.userdata_path),
            "stats": self.get_stats(),
        }

    def get_save_metadata(self) -> dict[str, Any]:
        self._require_open_save()
        mode = packer.detect_repacker(self.work_dir).mode
        return {
            "savePath": str(self.save_file_path) if self.save_file_path else "",
            "mode": mode,
            "defaultExtension": ".dat" if mode == "PS" else ".sl2" if mode == "PC" else "",
        }

    def get_save_target_info(self, output_file: str) -> dict[str, Any]:
        self._require_open_save()
        output_path = Path(output_file)
        mode = packer.detect_repacker(self.work_dir).mode
        info: dict[str, Any] = {
            "outputPath": str(output_path),
            "mode": mode,
            "steamIdMismatch": False,
            "currentSteamId": None,
            "targetSteamId": None,
        }

        if mode != "PC":
            return info

        dir_name = output_path.parent.name
        if not dir_name.isdigit() or len(dir_name) != 17:
            return info

        target_steam_id = int(dir_name)
        current_steam_id = int.from_bytes(packer.read_steam_id(self.work_dir), "little")
        info.update({
            "steamIdMismatch": target_steam_id != current_steam_id,
            "currentSteamId": current_steam_id,
            "targetSteamId": target_steam_id,
        })
        return info

    def save_as(self, output_file: str, resign_steam_id: bool = False) -> dict[str, Any]:
        self._require_character()
        output_path = Path(output_file)
        if not output_path.name:
            raise ValueError("Output save file path is required")

        self.save_current_character()
        resigned_steam_id = self._patch_steam_id_for_target(output_path) if resign_steam_id else False

        if self.config.auto_backup:
            with create_backup(output_path, self.backup_dir, self.config.max_backups):
                packer.repack(self.work_dir, output_path)
        else:
            packer.repack(self.work_dir, output_path)

        self.save_file_path = output_path
        self.config.last_file = str(output_path)
        return {
            "savePath": str(output_path),
            "mode": packer.detect_repacker(self.work_dir).mode,
            "backupDir": str(self.backup_dir),
            "resignedSteamId": resigned_steam_id,
            "selectedCharacter": self._current_character_state(),
        }

    def add_relic(self, relic_type: str = "normal", count: int = 1) -> dict[str, Any]:
        self._require_character()
        normalized_type = relic_type.lower()
        if normalized_type not in {"normal", "deep"}:
            raise ValueError("Relic type must be normal or deep")
        if count < 1 or count > 999:
            raise ValueError("Relic count must be between 1 and 999")

        last_ga_handle = 0
        for _ in range(count):
            _, last_ga_handle = self.inventory.add_relic_to_inventory(normalized_type)
        self.save_current_character()
        self._reparse_current_character()
        state = self._current_character_state()
        state["addedCount"] = count
        state["lastGaHandle"] = last_ga_handle
        return state

    def delete_relic(self, ga_handle: int) -> dict[str, Any]:
        self._require_character()
        if ga_handle not in self.inventory.relics:
            raise ValueError("Relic not found in inventory")

        self.inventory.remove_relic_from_inventory(ga_handle)
        self.save_current_character()
        self._reparse_current_character()
        return self._current_character_state()

    def delete_relics(self, ga_handles: list[int]) -> dict[str, Any]:
        self._require_character()
        handles = self._validate_relic_handles(ga_handles)
        deleted_count = 0
        failed_count = 0
        last_error = ""

        for ga_handle in handles:
            try:
                self.inventory.remove_relic_from_inventory(ga_handle)
                deleted_count += 1
            except Exception as exc:
                failed_count += 1
                last_error = str(exc)

        if deleted_count:
            self.save_current_character()
        self._reparse_current_character()

        message = f"Deleted {deleted_count} relics"
        if failed_count:
            message += f", {failed_count} failed"
            if failed_count == 1 and last_error:
                message += f": {last_error}"

        return {
            "deleted": deleted_count,
            "failed": failed_count,
            "message": message,
            "selectedCharacter": self._current_character_state(),
        }

    def toggle_favorite_relic(self, ga_handle: int) -> dict[str, Any]:
        self._require_character()
        if ga_handle not in self.inventory.relics:
            raise ValueError("Relic not found in inventory")

        favorite = self.inventory.toggle_favorite_mark(ga_handle)
        self.save_current_character()
        self._reparse_current_character()
        state = self._current_character_state()
        state["favorite"] = favorite
        return state

    def toggle_favorite_relics(self, ga_handles: list[int]) -> dict[str, Any]:
        self._require_character()
        handles = self._validate_relic_handles(ga_handles)
        for ga_handle in handles:
            self.inventory.toggle_favorite_mark(ga_handle)
        self.save_current_character()
        self._reparse_current_character()
        return {
            "message": f"Toggled favorite mark on {len(handles)} relics",
            "selectedCharacter": self._current_character_state(),
        }

    def copy_relic_effects(self, ga_handles: list[int]) -> dict[str, Any]:
        self._require_character()
        handles = self._validate_relic_handles(ga_handles)
        relics = [self.inventory.relics[ga_handle] for ga_handle in handles]
        relics.sort(key=lambda relic: relic.state.is_deep)
        sorted_handles = [relic.ga_handle for relic in relics]
        unique_names = [
            relic.state.name
            for relic in relics
            if relic.state.real_item_id in UNIQUENESS_IDS
        ]
        return {
            "effectsText": self.inventory.stringify_relic_effects(sorted_handles),
            "count": len(sorted_handles),
            "uniqueNames": unique_names,
        }

    def paste_relic_effects(
        self,
        ga_handles: list[int],
        effects_text: str,
    ) -> dict[str, Any]:
        self._require_character()
        handles = self._validate_relic_handles(ga_handles)
        relics = [self.inventory.relics[ga_handle] for ga_handle in handles]
        relics.sort(key=lambda relic: relic.state.is_deep)
        effects_lines = self.inventory.parse_effects(effects_text)
        if len(effects_lines) == 1:
            effects_lines = [effects_lines[0]] * len(relics)
        elif len(effects_lines) != len(relics):
            raise ValueError(
                "Copied effect count does not match selected relic count: "
                f"copied {len(effects_lines)}, selected {len(relics)}"
            )

        success_count = 0
        failed_count = 0
        last_error = ""
        for index, relic in enumerate(relics):
            try:
                self.inventory.modify_relic(relic.ga_handle, None, *effects_lines[index])
                success_count += 1
            except Exception as exc:
                failed_count += 1
                last_error = str(exc)

        if success_count:
            self.save_current_character()
            fix_result = self.mass_fix_relics()
            selected_character = fix_result["selectedCharacter"]
        else:
            selected_character = self._current_character_state()

        message = f"Effects pasted to {success_count} relics"
        if failed_count:
            message += f", {failed_count} failed"
            if failed_count == 1 and last_error:
                message += f": {last_error}"

        return {
            "pasted": success_count,
            "failed": failed_count,
            "message": message,
            "selectedCharacter": selected_character,
        }

    def reindex_relics(self, target_index: int, ga_handles: list[int]) -> dict[str, Any]:
        self._require_character()
        handles = self._validate_relic_handles(ga_handles)
        ordered_handles = list(reversed(handles))
        rank_order = sorted(
            [
                {
                    "gaHandle": ga_handle,
                    "acquisitionId": self.inventory.ga_to_acquisition_id.get(ga_handle, 999999),
                }
                for ga_handle in self.inventory.relic_gas
            ],
            key=lambda relic: relic["acquisitionId"],
        )
        bounded_target = min(max(0, target_index), len(rank_order))
        if bounded_target >= len(rank_order):
            new_acquisition_id = self.inventory.request_new_acquisition_id()
        else:
            new_acquisition_id = int(rank_order[bounded_target]["acquisitionId"])

        self.inventory.reindex_acquisition_id_at(new_acquisition_id, *ordered_handles)
        self.save_current_character()
        self._reparse_current_character()
        return {
            "message": f"Moved {len(handles)} relics after #{bounded_target}",
            "selectedCharacter": self._current_character_state(),
        }

    def delete_illegal_relics(self) -> dict[str, Any]:
        self._require_character()
        targets = list(self.inventory.illegal_gas)
        if not targets:
            return {
                "deleted": 0,
                "failed": 0,
                "message": "No illegal relics found",
                "selectedCharacter": self._current_character_state(),
            }

        deleted_count = 0
        failed_count = 0
        last_error = ""
        for ga_handle in targets:
            if ga_handle not in self.inventory.relics:
                continue
            try:
                self.inventory.remove_relic_from_inventory(ga_handle)
                deleted_count += 1
            except Exception as exc:
                failed_count += 1
                last_error = str(exc)

        if deleted_count:
            self.save_current_character()
        self._reparse_current_character()

        if failed_count:
            message = f"Deleted {deleted_count} illegal relics, {failed_count} failed"
            if failed_count == 1 and last_error:
                message += f": {last_error}"
        else:
            message = f"Successfully deleted {deleted_count} illegal relics"

        return {
            "deleted": deleted_count,
            "failed": failed_count,
            "message": message,
            "selectedCharacter": self._current_character_state(),
        }

    def mass_fix_relics(self) -> dict[str, Any]:
        self._require_character()
        fixable_illegal: list[dict[str, Any]] = []
        fixable_strict: list[dict[str, Any]] = []
        unfixable_relics: list[dict[str, Any]] = []

        for ga_handle in self.inventory.relic_gas:
            entry = self.inventory.relics[ga_handle]
            state = entry.state
            real_id = state.real_item_id
            effects = list(state.effects_and_curses)

            if real_id in UNIQUENESS_IDS:
                continue

            is_illegal = ga_handle in self.inventory.illegal_gas
            is_strict_invalid = ga_handle in self.inventory.strict_invalid_gas
            if not is_illegal and not is_strict_invalid:
                continue

            item_name = self._relic_name(real_id)
            if is_illegal:
                fix = self._find_mass_fix(ga_handle, real_id, effects, item_name, allow_fallback=True)
                if fix is not None:
                    fixable_illegal.append(fix)
                else:
                    unfixable_relics.append({
                        "itemId": real_id,
                        "name": item_name,
                        "issue": "illegal",
                        "reason": "No valid ID found",
                    })
                continue

            fix = self._find_mass_fix(ga_handle, real_id, effects, item_name, allow_fallback=False)
            if fix is not None:
                fixable_strict.append(fix)
            else:
                unfixable_relics.append({
                    "itemId": real_id,
                    "name": item_name,
                    "issue": "strict",
                    "reason": "No valid permutation found",
                })

        fixable_relics = fixable_illegal + fixable_strict
        if not fixable_relics:
            return {
                "fixed": 0,
                "failed": 0,
                "illegalFixed": 0,
                "strictFixed": 0,
                "fallbackFixed": 0,
                "unfixable": unfixable_relics,
                "message": "No fixable relics found",
                "selectedCharacter": self._current_character_state(),
            }

        fixed_count = 0
        failed_count = 0
        for fix in fixable_relics:
            try:
                new_effects = (
                    self.relic_checker.sort_effects(fix["effects"])
                    if fix["fallback"]
                    else fix["effects"]
                )
                self.inventory.modify_relic(fix["gaHandle"], fix["newId"], *new_effects)
                fixed_count += 1
            except Exception:
                failed_count += 1

        if fixed_count:
            self.save_current_character()
        self._reparse_current_character()

        fallback_count = sum(1 for fix in fixable_relics if fix["fallback"])
        message = f"Fixed {fixed_count} relics"
        if failed_count:
            message += f", {failed_count} failed"

        return {
            "fixed": fixed_count,
            "failed": failed_count,
            "illegalFixed": len(fixable_illegal),
            "strictFixed": len(fixable_strict),
            "fallbackFixed": fallback_count,
            "unfixable": unfixable_relics,
            "message": message,
            "selectedCharacter": self._current_character_state(),
        }

    def update_relic(
        self,
        ga_handle: int,
        relic_id: int | None = None,
        effects: list[int] | None = None,
    ) -> dict[str, Any]:
        self._require_character()
        if ga_handle not in self.inventory.relics:
            raise ValueError("Relic not found in inventory")

        updates: dict[str, int] = {}
        if relic_id is not None:
            self._validate_uint32(relic_id, "Relic item ID")
            current_id = self.inventory.relics[ga_handle].state.real_item_id
            current_is_deep = self.game_data.is_deep_relic(current_id)
            next_is_deep = self.game_data.is_deep_relic(relic_id)
            if current_is_deep != next_is_deep:
                current_type = "Deep" if current_is_deep else "Normal"
                next_type = "Deep" if next_is_deep else "Normal"
                raise ValueError(
                    f"Cannot change from a {current_type} relic to a {next_type} relic ID"
                )
            updates["relic_id"] = relic_id

        if effects is not None:
            if len(effects) != 6:
                raise ValueError("Relic effects must include 3 effects and 3 curses")
            normalized_effects = [
                0xFFFFFFFF if effect_id == -1 else effect_id
                for effect_id in effects
            ]
            for effect_id in normalized_effects:
                self._validate_uint32(effect_id, "Relic effect ID")
            updates.update({
                "effect_1": normalized_effects[0],
                "effect_2": normalized_effects[1],
                "effect_3": normalized_effects[2],
                "curse_1": normalized_effects[3],
                "curse_2": normalized_effects[4],
                "curse_3": normalized_effects[5],
            })

        if not updates:
            raise ValueError("No relic updates provided")

        self.inventory.modify_relic(ga_handle, **updates)
        self.save_current_character()
        self._reparse_current_character()
        return self._current_character_state()

    def prepare_relic_edit(self, relic_id: int, effects: list[int]) -> dict[str, Any]:
        self._validate_uint32(relic_id, "Relic item ID")
        if len(effects) != 6:
            raise ValueError("Relic effects must include 3 effects and 3 curses")

        normalized_effects = [self._normalize_edit_effect(effect_id) for effect_id in effects]
        sorted_effects = self.relic_checker.sort_effects(normalized_effects)
        suggested_id = self._find_valid_relic_id_for_effects(relic_id, sorted_effects)
        prepared_id = suggested_id if suggested_id is not None else relic_id
        invalid_reason = self.relic_checker.check_invalidity(prepared_id, sorted_effects)
        strict_invalid = self.relic_checker.is_strict_invalid(
            prepared_id,
            sorted_effects,
            invalid_reason,
        )

        return {
            "relicId": prepared_id,
            "relicName": self._relic_name(prepared_id),
            "effects": sorted_effects,
            "effectNames": [self._effect_name(effect_id) for effect_id in sorted_effects],
            "changedRelicId": prepared_id != relic_id,
            "invalidReason": getattr(invalid_reason, "name", str(invalid_reason)),
            "strictInvalid": strict_invalid,
        }

    def inspect_relic_edit(self, relic_id: int, effects: list[int]) -> dict[str, Any]:
        self._validate_uint32(relic_id, "Relic item ID")
        if len(effects) != 6:
            raise ValueError("Relic effects must include 3 effects and 3 curses")

        normalized_effects = [self._normalize_edit_effect(effect_id) for effect_id in effects]
        invalid_reason, invalid_index = self.relic_checker.check_invalidity(
            relic_id,
            normalized_effects,
            True,
        )
        strict_invalid = self.relic_checker.is_strict_invalid(
            relic_id,
            normalized_effects,
            invalid_reason,
        )
        strict_reason = (
            self.relic_checker.get_strict_invalid_reason(relic_id, normalized_effects)
            if strict_invalid
            else None
        )
        status, title, detail = self._relic_edit_status_text(
            relic_id,
            normalized_effects,
            invalid_reason,
            invalid_index,
            strict_reason,
        )
        effect_slots, curse_slots = self._safe_relic_slot_count(relic_id)

        return {
            "relicId": relic_id,
            "relicName": self._relic_name(relic_id),
            "effects": normalized_effects,
            "effectNames": [self._effect_name(effect_id) for effect_id in normalized_effects],
            "invalidReason": getattr(invalid_reason, "name", str(invalid_reason)),
            "invalidIndex": invalid_index,
            "strictInvalid": strict_invalid,
            "strictReason": strict_reason,
            "status": status,
            "title": title,
            "detail": detail,
            "color": self._relic_color_name(relic_id),
            "deep": self.game_data.is_deep_relic(relic_id),
            "effectSlots": effect_slots,
            "curseSlots": curse_slots,
            "debugLines": self._relic_edit_debug_lines(
                relic_id,
                normalized_effects,
                invalid_reason,
                invalid_index,
            ),
        }

    def change_relic_color(
        self,
        ga_handle: int,
        relic_id: int,
        effects: list[int],
        target_color: str,
    ) -> dict[str, Any]:
        self._require_character()
        if ga_handle not in self.inventory.relics:
            raise ValueError("Relic not found in inventory")
        self._validate_uint32(relic_id, "Relic item ID")
        if len(effects) != 6:
            raise ValueError("Relic effects must include 3 effects and 3 curses")

        color_names = {
            "red": (0, "Red"),
            "blue": (1, "Blue"),
            "yellow": (2, "Yellow"),
            "green": (3, "Green"),
        }
        color_entry = color_names.get(target_color.strip().lower())
        if color_entry is None:
            raise ValueError("Target relic color must be Red, Blue, Yellow, or Green")
        target_color_id, target_color_name = color_entry

        assigned_to = [
            self.game_data.character_names[hero_type - 1]
            for hero_type in self.inventory.get_relic_equipped_by(ga_handle)
            if 0 < hero_type <= len(self.game_data.character_names)
        ]
        if assigned_to:
            raise ValueError(
                "Cannot change relic color while it is assigned to: "
                + ", ".join(assigned_to)
            )

        normalized_effects = [self._normalize_edit_effect(effect_id) for effect_id in effects]
        current_color_id = (
            int(self.game_data._relic_table.loc[relic_id, "relicColor"])
            if relic_id in self.game_data._relic_table.index
            else None
        )
        if current_color_id == target_color_id:
            return {
                "relicId": relic_id,
                "relicName": self._relic_name(relic_id),
                "color": target_color_name,
                "effects": normalized_effects,
                "changedRelicId": False,
                "alreadyTarget": True,
            }

        suggested_id = self._find_valid_relic_id_for_effects(
            relic_id,
            normalized_effects,
            target_color_id,
        )
        if suggested_id is None:
            used_effects = sum(1 for effect_id in normalized_effects[:3] if effect_id not in EMPTY_EFFECT_IDS)
            used_curses = sum(1 for effect_id in normalized_effects[3:] if effect_id not in EMPTY_EFFECT_IDS)
            raise ValueError(
                f"Could not find a valid {target_color_name} relic for "
                f"{used_effects} effect slot(s) and {used_curses} curse slot(s)"
            )

        return {
            "relicId": suggested_id,
            "relicName": self._relic_name(suggested_id),
            "color": target_color_name,
            "effects": normalized_effects,
            "changedRelicId": suggested_id != relic_id,
            "alreadyTarget": False,
        }

    def list_relic_edit_options(
        self,
        relic_id: int,
        safe_mode: bool = True,
    ) -> list[dict[str, Any]]:
        self._validate_uint32(relic_id, "Relic item ID")
        current_is_deep = self.game_data.is_deep_relic(relic_id)
        if safe_mode:
            relic_ids = self.game_data.get_safe_relic_ids()
        else:
            relic_ids = [int(candidate_id) for candidate_id in self.game_data.relics.keys()]

        rows: list[dict[str, Any]] = []
        for candidate_id in sorted(set(relic_ids)):
            relic = self.game_data.relics.get(candidate_id)
            if relic is None:
                continue
            candidate_is_deep = self.game_data.is_deep_relic(candidate_id)
            if candidate_is_deep != current_is_deep:
                continue
            effect_slots, curse_slots = self.game_data.get_relic_slot_count(candidate_id)
            rows.append({
                "id": candidate_id,
                "name": relic.name,
                "color": relic.color,
                "deep": candidate_is_deep,
                "effectSlots": effect_slots,
                "curseSlots": curse_slots,
            })
        return rows

    def list_effect_edit_options(
        self,
        relic_id: int,
        slot_index: int,
        effects: list[int],
        safe_mode: bool = True,
    ) -> list[dict[str, Any]]:
        self._validate_uint32(relic_id, "Relic item ID")
        if slot_index < 0 or slot_index >= 6:
            raise ValueError("Invalid relic effect slot index")
        if len(effects) != 6:
            raise ValueError("Relic effects must include 3 effects and 3 curses")

        normalized_effects = [self._normalize_edit_effect(effect_id) for effect_id in effects]
        is_curse_slot = slot_index >= 3
        if safe_mode:
            if is_curse_slot:
                pool_type = "curse"
            elif self.game_data.is_deep_relic(relic_id):
                pool_type = "deep"
            else:
                pool_type = "normal"
            effect_ids = list(self.game_data.get_rollable_effects(pool_type))
        else:
            effect_ids = [int(effect_id) for effect_id in self.game_data.effects.keys()]

        warned_ids = self._conflicting_effect_option_ids(
            normalized_effects,
            slot_index,
            effect_ids,
        )
        option_ids = [0xFFFFFFFF]
        option_ids.extend(effect_id for effect_id in sorted(set(effect_ids)) if effect_id != 0xFFFFFFFF)

        return [
            {
                "id": effect_id,
                "name": self._effect_name(effect_id),
                "warning": effect_id in warned_ids,
                "needsCurse": self.game_data.effect_needs_curse(effect_id),
            }
            for effect_id in option_ids
        ]

    def export_relics_excel(self, output_file: str) -> dict[str, Any]:
        self._require_character()
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        output_path = Path(output_file)
        if not output_path.name:
            raise ValueError("Output Excel file path is required")
        if not self.inventory.relic_gas:
            raise RuntimeError("No relics found in inventory")

        wb = Workbook()
        ws = wb.active
        ws.title = "Relics"
        headers = [
            "Item ID",
            "Relic Name",
            "Relic Color",
            "Effect 1 (ID)",
            "Effect 1 (Name)",
            "Effect 2 (ID)",
            "Effect 2 (Name)",
            "Effect 3 (ID)",
            "Effect 3 (Name)",
            "Curse 1 (ID)",
            "Curse 1 (Name)",
            "Curse 2 (ID)",
            "Curse 2 (Name)",
            "Curse 3 (ID)",
            "Curse 3 (Name)",
        ]
        ws.append(headers)

        for relic_ga in self.inventory.relic_gas:
            state = self.inventory.relics[relic_ga].state
            real_id = state.real_item_id
            if real_id == 0x00FFFFFF:
                continue
            item = self.game_data.relics.get(real_id)
            effects = list(state.effects_and_curses)
            row = [
                real_id,
                item.name if item else f"UnknownRelic({real_id})",
                item.color if item else "Unknown",
                effects[0],
                self._effect_name_for_excel(effects[0]),
                effects[1],
                self._effect_name_for_excel(effects[1]),
                effects[2],
                self._effect_name_for_excel(effects[2]),
                effects[3],
                self._effect_name_for_excel(effects[3]),
                effects[4],
                self._effect_name_for_excel(effects[4]),
                effects[5],
                self._effect_name_for_excel(effects[5]),
            ]
            ws.append(row)

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                value = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(value))
            ws.column_dimensions[col_letter].width = max_len + 2

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return {"path": str(output_path), "count": len(self.inventory.relic_gas)}

    def import_relics_excel(self, input_file: str) -> dict[str, Any]:
        self._require_character()
        from openpyxl import load_workbook

        input_path = Path(input_file)
        if not input_path.is_file():
            raise FileNotFoundError(f"Excel file not found: {input_path}")

        try:
            wb = load_workbook(input_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
        except Exception as exc:
            raise ValueError(f"Failed to read Excel: {exc}") from exc

        excel_sigs: dict[tuple[int, int, int, int, int, int, int], int] = {}
        skipped_unique_count = 0
        for row in rows:
            if not row or row[0] is None:
                continue
            real_id = int(row[0])
            if real_id in UNIQUENESS_IDS:
                skipped_unique_count += 1
                continue
            sig = (
                real_id,
                self._norm_effect(row[3]),
                self._norm_effect(row[5]),
                self._norm_effect(row[7]),
                self._norm_effect(row[9]),
                self._norm_effect(row[11]),
                self._norm_effect(row[13]),
            )
            excel_sigs[sig] = excel_sigs.get(sig, 0) + 1

        if not excel_sigs and skipped_unique_count == 0:
            raise ValueError("No valid relic rows found in Excel file")

        inv_sigs: dict[tuple[int, int, int, int, int, int, int], int] = {}
        for ga_handle in self.inventory.relic_gas:
            state = self.inventory.relics[ga_handle].state
            effects = list(state.effects_and_curses)
            sig = (
                state.real_item_id,
                self._norm_effect(effects[0]),
                self._norm_effect(effects[1]),
                self._norm_effect(effects[2]),
                self._norm_effect(effects[3]),
                self._norm_effect(effects[4]),
                self._norm_effect(effects[5]),
            )
            inv_sigs[sig] = inv_sigs.get(sig, 0) + 1

        total_to_add = sum(
            max(0, excel_count - inv_sigs.get(sig, 0))
            for sig, excel_count in excel_sigs.items()
        )
        if total_to_add == 0:
            already_had = sum(
                min(excel_sigs[sig], inv_sigs[sig])
                for sig in set(excel_sigs) & set(inv_sigs)
            )
            return {
                "added": 0,
                "failed": 0,
                "existing": already_had,
                "skippedUnique": skipped_unique_count,
                "selectedCharacter": self._current_character_state(),
            }

        added_count = 0
        failed_count = 0
        for sig, excel_count in excel_sigs.items():
            inv_count = inv_sigs.get(sig, 0)
            if excel_count <= inv_count:
                continue
            real_id = sig[0]
            effects = list(sig[1:])
            for _ in range(excel_count - inv_count):
                try:
                    _, new_ga = self.inventory.add_relic_to_inventory()
                    self.inventory.modify_relic(new_ga, real_id, *effects)
                    added_count += 1
                except Exception:
                    failed_count += 1

        self.save_current_character()
        self._reparse_current_character()
        already_had = sum(
            min(excel_sigs[sig], inv_sigs[sig])
            for sig in set(excel_sigs) & set(inv_sigs)
        )
        return {
            "added": added_count,
            "failed": failed_count,
            "existing": already_had,
            "skippedUnique": skipped_unique_count,
            "selectedCharacter": self._current_character_state(),
        }

    def list_relics(self) -> list[dict[str, Any]]:
        self._require_character()
        rows: list[dict[str, Any]] = []
        for order, ga_handle in enumerate(self.inventory.relic_gas, start=1):
            entry = self.inventory.relics[ga_handle]
            state = entry.state
            relic_id = state.real_item_id
            relic = self.game_data.relics.get(relic_id)
            effects = list(state.effects_and_curses)
            equipped_hero_types = self.inventory.get_relic_equipped_by(ga_handle)
            equipped_by = [
                self.game_data.character_names[hero_type - 1]
                for hero_type in equipped_hero_types
                if 1 <= hero_type <= len(self.game_data.character_names)
            ]

            rows.append({
                "order": order,
                "gaHandle": ga_handle,
                "gaHandleHex": f"0x{ga_handle:08X}",
                "itemId": relic_id,
                "name": relic.name if relic else f"Unknown ({relic_id})",
                "color": relic.color if relic else "Unknown",
                "deep": bool(relic.is_deep()) if relic else False,
                "favorite": bool(entry.is_favorite),
                "new": bool(entry.is_new),
                "equippedBy": equipped_by,
                "equippedByText": ", ".join(equipped_by) if equipped_by else "-",
                "effectIds": effects,
                "effectNames": [self._effect_name(effect_id) for effect_id in effects],
                "illegal": ga_handle in self.inventory.illegal_gas,
                "curseIllegal": ga_handle in self.inventory.curse_illegal_gas,
                "strictInvalid": ga_handle in self.inventory.strict_invalid_gas,
                "unique": relic_id in UNIQUENESS_IDS,
            })
        return rows

    def list_vessels(self, hero_type: int = 1) -> list[dict[str, Any]]:
        self._require_character()
        if hero_type not in self.loadouts.heroes:
            return []

        hero = self.loadouts.heroes[hero_type]
        groups: list[dict[str, Any]] = []
        for vessel_index, vessel in enumerate(hero.vessels):
            vessel_id = int(vessel["vessel_id"])
            vessel_meta = self.game_data.vessels.get(vessel_id)
            slot_colors = (
                list(vessel_meta.slot_colors)
                if vessel_meta
                else ["Unknown", "Unknown", "Unknown", "Unknown", "Unknown", "Unknown"]
            )
            unlocked = is_vessel_available(vessel_id)
            rows = [
                self._serialize_vessel_slot(slot_index, ga_handle, slot_colors)
                for slot_index, ga_handle in enumerate(vessel["relics"])
            ]
            groups.append({
                "index": vessel_index,
                "vesselId": vessel_id,
                "name": vessel_meta.name if vessel_meta else f"Vessel {vessel_index + 1}",
                "heroType": hero_type,
                "unlocked": unlocked,
                "unlockFlag": int(vessel_meta.unlock_flag) if vessel_meta else -1,
                "status": (
                    f"已解锁（{sum(1 for row in rows if not row['empty'])}/6 遗物）"
                    if unlocked
                    else f"未解锁（Unlock Flag: {int(vessel_meta.unlock_flag) if vessel_meta else -1}）"
                ),
                "rows": rows,
            })
        return groups

    def list_presets(self, hero_type: int = 1) -> list[dict[str, Any]]:
        self._require_character()
        if hero_type not in self.loadouts.heroes:
            return []

        hero = self.loadouts.heroes[hero_type]
        return [
            self._serialize_preset(hero_type, hero_preset_index, preset)
            for hero_preset_index, preset in enumerate(hero.presets)
        ]

    def list_heroes(self) -> list[dict[str, Any]]:
        return [
            {
                "heroType": index + 1,
                "name": name,
            }
            for index, name in enumerate(self.game_data.character_names[:10])
        ]

    def list_vessel_relic_options(
        self,
        hero_type: int,
        vessel_id: int,
        slot_index: int,
    ) -> list[dict[str, Any]]:
        self._require_character()
        if slot_index < 0 or slot_index >= 6:
            raise ValueError("Invalid relic slot index")

        vessel_meta = self.game_data.vessels.get(vessel_id)
        if not vessel_meta:
            raise ValueError("Vessel not found")

        required_color = (
            vessel_meta.slot_colors[slot_index]
            if slot_index < len(vessel_meta.slot_colors)
            else "Unknown"
        )
        requires_deep = slot_index >= 3
        options = [{
            "gaHandle": 0,
            "itemId": None,
            "name": "(Empty)",
            "color": required_color,
            "deep": requires_deep,
            "equippedByText": "-",
            "effectNames": ["-", "-", "-"],
        }]

        for ga_handle in self.inventory.relic_gas:
            entry = self.inventory.relics[ga_handle]
            state = entry.state
            relic = self.game_data.relics.get(state.real_item_id)
            if relic is None:
                continue
            if bool(relic.is_deep()) != requires_deep:
                continue
            if required_color != "White" and relic.color != required_color:
                continue

            equipped_hero_types = self.inventory.get_relic_equipped_by(ga_handle)
            equipped_by = [
                self.game_data.character_names[hero - 1]
                for hero in equipped_hero_types
                if 1 <= hero <= len(self.game_data.character_names)
            ]
            effects = list(state.effects_and_curses)
            options.append({
                "gaHandle": ga_handle,
                "itemId": state.real_item_id,
                "name": relic.name,
                "color": relic.color,
                "deep": bool(relic.is_deep()),
                "equippedByText": ", ".join(equipped_by) if equipped_by else "-",
                "effectNames": [self._effect_name(effect_id) for effect_id in effects[:3]],
            })

        return options

    def replace_vessel_relic(
        self,
        hero_type: int,
        vessel_id: int,
        slot_index: int,
        ga_handle: int,
    ) -> dict[str, Any]:
        self._require_character()
        if slot_index < 0 or slot_index >= 6:
            raise ValueError("Invalid relic slot index")
        if ga_handle != 0 and ga_handle not in self.inventory.relics:
            raise ValueError("Relic not found in inventory")

        self.loadouts.replace_vessel_relic(hero_type, vessel_id, slot_index, ga_handle)
        self.save_current_character()
        self._reparse_current_character()
        return self._current_character_state(hero_type=hero_type)

    def replace_preset_relic(
        self,
        hero_type: int,
        preset_index: int,
        slot_index: int,
        ga_handle: int,
    ) -> dict[str, Any]:
        self._require_character()
        if slot_index < 0 or slot_index >= 6:
            raise ValueError("Invalid relic slot index")
        if preset_index < 0 or preset_index >= len(self.loadouts.all_presets):
            raise ValueError("Invalid preset index")
        if ga_handle != 0 and ga_handle not in self.inventory.relics:
            raise ValueError("Relic not found in inventory")

        preset = self.loadouts.all_presets[preset_index]
        if int(preset.get("hero_type", hero_type)) != hero_type:
            raise ValueError("Preset does not belong to selected hero")

        self.loadouts.replace_preset_relic(
            hero_type,
            slot_index,
            ga_handle,
            preset_index=preset_index,
        )
        self.save_current_character()
        self._reparse_current_character()
        return self._current_character_state(hero_type=hero_type)

    def save_vessel_as_preset(
        self,
        hero_type: int,
        vessel_id: int,
        name: str,
    ) -> dict[str, Any]:
        self._require_character()
        preset_name = self._validate_preset_name(name)
        if hero_type not in self.loadouts.heroes:
            raise ValueError("Hero not found")

        relics = None
        for vessel in self.loadouts.heroes[hero_type].vessels:
            if int(vessel["vessel_id"]) == vessel_id:
                relics = list(vessel["relics"])
                break
        if relics is None:
            raise ValueError("Vessel not found")

        self.loadouts.push_preset(hero_type, vessel_id, relics, preset_name)
        self.save_current_character()
        self._reparse_current_character()
        return self._current_character_state(hero_type=hero_type)

    def equip_preset(self, hero_type: int, preset_index: int) -> dict[str, Any]:
        self._require_character()
        self.loadouts.equip_preset(hero_type, preset_index)
        self.save_current_character()
        self._reparse_current_character()
        return self._current_character_state(hero_type=hero_type)

    def delete_preset(self, hero_type: int, preset_index: int) -> dict[str, Any]:
        self._require_character()
        self.loadouts.remove_preset(preset_index)
        self.save_current_character()
        self._reparse_current_character()
        return self._current_character_state(hero_type=hero_type)

    def rename_preset(self, hero_type: int, preset_index: int, name: str) -> dict[str, Any]:
        self._require_character()
        preset_name = self._validate_preset_name(name)
        self.loadouts.rename_preset(preset_index, preset_name)
        self.save_current_character()
        self._reparse_current_character()
        return self._current_character_state(hero_type=hero_type)

    def export_loadout(self, hero_type: int, output_file: str) -> dict[str, Any]:
        self._require_character()
        output_path = Path(output_file)
        if not output_path.name:
            raise ValueError("Output loadout file path is required")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.loadouts.export_hero_loadout(hero_type, str(output_path))
        return {
            "path": str(output_path),
            "heroType": hero_type,
        }

    def preview_import_loadout(self, input_file: str) -> dict[str, Any]:
        self._require_character()
        input_path = Path(input_file)
        if not input_path.is_file():
            raise FileNotFoundError(f"Loadout file not found: {input_path}")

        self._cancel_pending_loadout_import()
        assert globals.data is not None
        self.pending_loadout_import_snapshot = bytearray(globals.data)
        import_generator = self.loadouts.import_hero_loadout(str(input_path))
        try:
            preview = next(import_generator)
        except StopIteration as exc:
            self._cancel_pending_loadout_import()
            value = exc.value
            if isinstance(value, list):
                raise RuntimeError("\n".join(str(item) for item in value))
            raise RuntimeError("Loadout import did not return preview data")
        if not isinstance(preview, dict):
            self._cancel_pending_loadout_import()
            raise RuntimeError("Invalid loadout preview data")

        self.pending_loadout_import = import_generator
        return self._serialize_loadout_import_preview(str(input_path), preview)

    def cancel_import_loadout(self) -> dict[str, Any]:
        self._require_character()
        restored = self._cancel_pending_loadout_import()
        return {
            "restored": restored,
            "selectedCharacter": self._current_character_state(),
        }

    def apply_import_loadout(
        self,
        vessel_indices: list[int],
        preset_indices: list[int],
    ) -> dict[str, Any]:
        self._require_character()
        if self.pending_loadout_import is None:
            raise RuntimeError("No pending loadout import")

        try:
            messages = self.pending_loadout_import.send((vessel_indices, preset_indices))
        except StopIteration as exc:
            messages = exc.value
        finally:
            self.pending_loadout_import = None
            self.pending_loadout_import_snapshot = None

        if messages is None:
            messages = []
        if not isinstance(messages, list):
            messages = [str(messages)]

        self.save_current_character()
        self._reparse_current_character()
        return {
            "messages": [str(message) for message in messages],
            "selectedCharacter": self._current_character_state(),
        }

    def _cancel_pending_loadout_import(self) -> bool:
        restored = False
        if self.pending_loadout_import_snapshot is not None:
            globals.data = bytearray(self.pending_loadout_import_snapshot)
            self._reparse_current_character()
            restored = True
        self.pending_loadout_import = None
        self.pending_loadout_import_snapshot = None
        return restored

    def _current_character_state(self, hero_type: int = 1) -> dict[str, Any]:
        self._require_character()
        if self.selected_character_index is None:
            raise RuntimeError("No character is selected")
        name, path = self.characters[self.selected_character_index]
        return {
            "index": self.selected_character_index,
            "name": name,
            "path": str(path),
            "stats": self.get_stats(),
            "characters": self._serialize_characters(),
            "heroes": self.list_heroes(),
            "relics": self.list_relics(),
            "vessels": self.list_vessels(hero_type=hero_type),
            "presets": self.list_presets(hero_type=hero_type),
        }

    def _serialize_settings(self) -> dict[str, Any]:
        return {
            "configPath": str(CONFIG_FILE),
            "lastFile": self.config.last_file,
            "lastCharIndex": int(self.config.last_char_index),
            "language": self.config.language,
            "languageName": LANGUAGE_MAP.get(self.config.language, self.config.language),
            "languages": [
                {"code": code, "name": name}
                for code, name in LANGUAGE_MAP.items()
            ],
            "theme": self.config.theme,
            "reduceMessagePop": bool(self.config.reduce_message_pop),
            "autoBackup": bool(self.config.auto_backup),
            "maxBackups": int(self.config.max_backups),
        }

    def _reparse_current_character(self) -> None:
        self.inventory.parse()
        self.loadouts.parse()
        self.inventory.set_illegal_relics()

    def _serialize_vessel_slot(
        self,
        slot_index: int,
        ga_handle: int,
        slot_colors: list[str],
    ) -> dict[str, Any]:
        required_color = slot_colors[slot_index] if slot_index < len(slot_colors) else "Unknown"
        slot_type = "Deep" if slot_index >= 3 else "Normal"
        base = {
            "slot": slot_index + 1,
            "type": slot_type,
            "requiredColor": required_color,
            "gaHandle": ga_handle,
            "empty": ga_handle == 0,
        }
        if ga_handle == 0 or ga_handle not in self.inventory.relics:
            return {
                **base,
                "itemId": None,
                "name": "(Empty)",
                "color": required_color,
                "deep": slot_type == "Deep",
                "effectIds": [],
                "effectNames": ["-", "-", "-"],
            }

        entry = self.inventory.relics[ga_handle]
        state = entry.state
        relic_id = state.real_item_id
        relic = self.game_data.relics.get(relic_id)
        effects = list(state.effects_and_curses)
        return {
            **base,
            "itemId": relic_id,
            "name": relic.name if relic else f"Unknown ({relic_id})",
            "color": relic.color if relic else required_color,
            "deep": bool(relic.is_deep()) if relic else slot_type == "Deep",
            "effectIds": effects,
            "effectNames": [self._effect_name(effect_id) for effect_id in effects[:3]],
        }

    def _serialize_preset(
        self,
        hero_type: int,
        hero_preset_index: int,
        preset: dict[str, Any],
    ) -> dict[str, Any]:
        vessel_id = int(preset["vessel_id"])
        vessel_meta = self.game_data.vessels.get(vessel_id)
        slot_colors = (
            list(vessel_meta.slot_colors)
            if vessel_meta
            else ["Unknown", "Unknown", "Unknown", "Unknown", "Unknown", "Unknown"]
        )
        rows = [
            self._serialize_vessel_slot(slot_index, ga_handle, slot_colors)
            for slot_index, ga_handle in enumerate(preset.get("relics", []))
        ]
        return {
            "index": int(preset["index"]),
            "heroPresetIndex": hero_preset_index,
            "heroType": int(preset.get("hero_type", hero_type)),
            "name": str(preset.get("name") or f"Preset {hero_preset_index + 1}"),
            "vesselId": vessel_id,
            "vesselName": vessel_meta.name if vessel_meta else f"Vessel {vessel_id}",
            "relicCount": sum(1 for row in rows if not row["empty"]),
            "equipped": self.loadouts.heroes[hero_type].cur_preset_idx == int(preset["index"]),
            "rows": rows,
        }

    def _serialize_characters(self) -> list[dict[str, Any]]:
        return [
            {
                "index": index,
                "name": name,
                "path": str(path),
                "selected": index == self.selected_character_index,
            }
            for index, (name, path) in enumerate(self.characters)
        ]

    def _serialize_import_characters(self) -> list[dict[str, Any]]:
        return [
            {
                "index": index,
                "name": name,
                "path": str(path),
                "selected": False,
            }
            for index, (name, path) in enumerate(self.import_characters)
        ]

    def _serialize_loadout_import_preview(
        self,
        input_file: str,
        preview: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "inputFile": input_file,
            "vessels": [
                self._serialize_loadout_import_item(index, vessel, "vessel")
                for index, vessel in enumerate(preview.get("vessels", []))
            ],
            "presets": [
                self._serialize_loadout_import_item(index, preset, "preset")
                for index, preset in enumerate(preview.get("presets", []))
            ],
        }

    def _serialize_loadout_import_item(
        self,
        index: int,
        item: dict[str, Any],
        item_type: str,
    ) -> dict[str, Any]:
        vessel_id = int(item["vessel_id"])
        vessel_meta = self.game_data.vessels.get(vessel_id)
        relics = list(item.get("relics", []))
        return {
            "index": index,
            "type": item_type,
            "name": item.get("name") or (vessel_meta.name if vessel_meta else f"Vessel {vessel_id}"),
            "vesselId": vessel_id,
            "vesselName": vessel_meta.name if vessel_meta else f"Vessel {vessel_id}",
            "relicCount": sum(1 for ga_handle in relics if ga_handle),
            "unlocked": is_vessel_available(vessel_id),
        }

    def _refresh_characters_preserving_selection(self) -> None:
        if self.userdata_path is None:
            self.characters = self._name_to_path(self.work_dir)
            return

        selected_path = self.userdata_path.resolve()
        self.characters = self._name_to_path(self.work_dir)
        self.selected_character_index = None
        for index, (_name, path) in enumerate(self.characters):
            if path.resolve() == selected_path:
                self.selected_character_index = index
                break

    def _effect_name(self, effect_id: int) -> str:
        if effect_id in EMPTY_EFFECT_IDS:
            return "-"
        effect = self.game_data.effects.get(effect_id)
        return effect.name if effect else f"Unknown ({effect_id})"

    def _relic_color_name(self, relic_id: int) -> str:
        relic = self.game_data.relics.get(relic_id)
        if relic is not None:
            return relic.color
        return "Unknown"

    def _safe_relic_slot_count(self, relic_id: int) -> tuple[int, int]:
        try:
            return self.game_data.get_relic_slot_count(relic_id)
        except Exception:
            return 0, 0

    def _relic_edit_status_text(
        self,
        relic_id: int,
        effects: list[int],
        invalid_reason: InvalidReason,
        invalid_index: int,
        strict_reason: str | None,
    ) -> tuple[str, str, str]:
        if invalid_reason == InvalidReason.NONE:
            if strict_reason:
                return (
                    "strictInvalid",
                    "STRICT INVALID",
                    strict_reason
                    + " This is technically valid but may have 0% drop-weight effects.",
                )
            return (
                "valid",
                "VALID",
                "This relic configuration is legal and should work in-game.",
            )

        reason_messages = {
            InvalidReason.IN_ILLEGAL_RANGE: "Relic ID is in an illegal or reserved range.",
            InvalidReason.INVALID_ITEM: f"Relic ID {relic_id} is not a valid relic.",
            InvalidReason.EFF_MUST_EMPTY: "An effect slot should be empty but has a value.",
            InvalidReason.EFF_NOT_ASSIGNED: "An effect slot must be assigned but is empty.",
            InvalidReason.EFF_NOT_IN_ROLLABLE_POOL: "One or more effects cannot roll on this relic.",
            InvalidReason.EFF_CONFLICT: "Two effects have the same conflict ID and cannot be combined.",
            InvalidReason.CURSE_MUST_EMPTY: "A curse slot should be empty but has a value.",
            InvalidReason.CURSE_REQUIRED_BY_EFFECT: "An effect requires a matching curse but none is assigned.",
            InvalidReason.CURSE_NOT_IN_ROLLABLE_POOL: "One or more curses cannot roll on this relic.",
            InvalidReason.CURSE_CONFLICT: "Two curses have the same conflict ID and cannot be combined.",
            InvalidReason.CURSES_NOT_ENOUGH: "Not enough curses are provided for effects that require curses.",
            InvalidReason.CURSE_SLOT_UNNECESSARY: "There are more curse slots than needed for the assigned effects.",
            InvalidReason.EFFS_NOT_SORTED: "Effects are valid but not in the required sorted order.",
            InvalidReason.VALIDATION_ERROR: "The relic could not be validated with the current data.",
        }
        detail = reason_messages.get(invalid_reason, f"Invalid relic configuration: {invalid_reason}")
        if 0 <= invalid_index < len(effects):
            slot_name = "Effect" if invalid_index < 3 else "Curse"
            slot_number = invalid_index + 1 if invalid_index < 3 else invalid_index - 2
            detail += f" First problematic slot: {slot_name} {slot_number}."
        return "illegal", "ILLEGAL", detail

    def _relic_edit_debug_lines(
        self,
        relic_id: int,
        effects: list[int],
        invalid_reason: InvalidReason,
        invalid_index: int,
    ) -> list[str]:
        lines = [
            f"Relic ID: {relic_id}",
            f"Effects: {effects[:3]}",
            f"Curses:  {effects[3:]}",
            "",
        ]
        try:
            pools = self.game_data.relics[relic_id].effect_slots
            lines.append(f"Pools (eff1,eff2,eff3,curse1,curse2,curse3): {pools}")
        except KeyError:
            lines.append(f"Pools: Relic ID {relic_id} not found in table")
            return lines

        invalid_suffix = "" if invalid_index == -1 else f" at effect index {invalid_index + 1}"
        curse_suffix = "" if invalid_index == -1 else f" at curse index {invalid_index - 2}"
        lines.extend([
            "",
            f"invalid_reason(): {invalid_reason.name}{invalid_suffix}",
            f"is_curse_illegal(): {is_curse_invalid(invalid_reason)}{curse_suffix}",
            "",
            "--- Effect Analysis ---",
        ])

        effect_slot_count = sum(1 for pool in pools[:3] if pool != -1)
        lines.append(
            f"Effect slots: {effect_slot_count} "
            f"({'single-effect' if effect_slot_count <= 1 else 'multi-effect'} relic)"
        )
        if effect_slot_count <= 1:
            lines.append("Single-effect relics do not require curses")
        else:
            curse_required_count = 0
            for index, effect_id in enumerate(effects[:3]):
                if effect_id in EMPTY_EFFECT_IDS:
                    lines.append(f"Effect {index}: {effect_id} (empty)")
                    continue
                needs_curse = self.game_data.effect_needs_curse(effect_id)
                effect_pools = self.game_data.get_effect_pools(effect_id)
                if needs_curse:
                    curse_required_count += 1
                lines.append(
                    f"Effect {index}: {effect_id} -> "
                    f"needs_curse={needs_curse}, pools={effect_pools}"
                )

            curses_provided = sum(
                1 for curse_id in effects[3:] if curse_id not in EMPTY_EFFECT_IDS
            )
            lines.extend([
                "",
                f"Effects needing curses: {curse_required_count}, Curses provided: {curses_provided}",
            ])
            if curse_required_count > curses_provided:
                lines.append("NOT ENOUGH CURSES for effects that need them")

        lines.extend(["", "--- Sequence Check Details ---"])
        if effect_slot_count <= 1:
            lines.append("(Single-effect relic - curse checks skipped)")

        for sequence in ([0, 1, 2], [0, 2, 1], [1, 0, 2], [1, 2, 0], [2, 0, 1], [2, 1, 0]):
            current_effects = [effects[index] for index in sequence]
            current_curses = [effects[index + 3] for index in sequence]
            valid = True
            issues: list[str] = []

            for slot_index in range(3):
                effect_id = current_effects[slot_index]
                curse_id = current_curses[slot_index]
                effect_pool = pools[slot_index]
                curse_pool = pools[slot_index + 3]

                if effect_pool == -1:
                    if effect_id != 0xFFFFFFFF:
                        valid = False
                        issues.append(f"slot{slot_index}: pool=-1 but eff={effect_id}")
                else:
                    pool_effects = self.game_data.get_pool_effects_strict(effect_pool)
                    if effect_id not in pool_effects:
                        valid = False
                        issues.append(f"slot{slot_index}: eff {effect_id} not in pool {effect_pool}")

                if effect_slot_count <= 1:
                    continue

                effect_needs_curse = self.game_data.effect_needs_curse(effect_id)
                if effect_needs_curse:
                    if curse_pool == -1:
                        valid = False
                        issues.append(f"curse{slot_index}: effect needs curse but slot has no curse_pool")
                    elif curse_id in EMPTY_EFFECT_IDS:
                        valid = False
                        issues.append(f"curse{slot_index}: effect requires curse, but empty")
                    elif curse_id not in self.game_data.get_pool_effects_strict(curse_pool):
                        valid = False
                        issues.append(f"curse{slot_index}: {curse_id} not in pool {curse_pool}")
                elif curse_pool != -1:
                    if (
                        curse_id not in EMPTY_EFFECT_IDS
                        and curse_id not in self.game_data.get_pool_effects_strict(curse_pool)
                    ):
                        valid = False
                        issues.append(f"curse{slot_index}: {curse_id} not in pool {curse_pool}")
                elif curse_id not in EMPTY_EFFECT_IDS:
                    valid = False
                    issues.append(f"curse{slot_index}: slot does not support curse but curse={curse_id}")

            lines.append(f"Seq {sequence}: {'VALID' if valid else 'invalid'}")
            lines.extend(f"  - {issue}" for issue in issues)

        return lines

    def _effect_name_for_excel(self, effect_id: int) -> str:
        if effect_id in EMPTY_EFFECT_IDS:
            return "None"
        effect = self.game_data.effects.get(effect_id)
        return effect.name if effect else f"UnknownEffect({effect_id})"

    @staticmethod
    def _norm_effect(value: Any) -> int:
        if value is None:
            return 0xFFFFFFFF
        effect_id = int(value)
        if effect_id in EMPTY_EFFECT_IDS:
            return 0xFFFFFFFF
        return effect_id

    def _normalize_edit_effect(self, value: Any) -> int:
        effect_id = int(value)
        if effect_id == -1:
            return 0xFFFFFFFF
        self._validate_uint32(effect_id, "Relic effect ID")
        if effect_id in EMPTY_EFFECT_IDS or effect_id not in self.game_data.effects:
            return 0xFFFFFFFF
        return effect_id

    def _conflicting_effect_option_ids(
        self,
        effects: list[int],
        slot_index: int,
        option_ids: list[int],
    ) -> set[int]:
        comparison_range = range(3, 6) if slot_index >= 3 else range(3)
        existing_conflict_ids: set[int] = set()
        for current_index in comparison_range:
            if current_index == slot_index:
                continue
            effect_id = effects[current_index]
            effect = self.game_data.effects.get(effect_id)
            if effect is not None and effect.conflict_id != -1:
                existing_conflict_ids.add(effect.conflict_id)

        if not existing_conflict_ids:
            return set()

        warned_ids: set[int] = set()
        for option_id in option_ids:
            effect = self.game_data.effects.get(option_id)
            if effect is not None and effect.conflict_id in existing_conflict_ids:
                warned_ids.add(option_id)
        return warned_ids

    def _require_character(self) -> None:
        if globals.data is None or self.userdata_path is None:
            raise RuntimeError("No character is loaded")

    def _require_open_save(self) -> None:
        if self.save_file_path is None or not self.work_dir.exists():
            raise RuntimeError("No save file is open")

    def _patch_steam_id_for_target(self, output_path: Path) -> bool:
        info = self.get_save_target_info(str(output_path))
        if not info["steamIdMismatch"] or info["targetSteamId"] is None:
            return False

        target_steam_id_bytes = int(info["targetSteamId"]).to_bytes(8, byteorder="little")
        for index in range(11):
            userdata_file = self.work_dir / f"USERDATA_{index}"
            if userdata_file.exists():
                packer.patch_steam_id(userdata_file, target_steam_id_bytes)
        return True

    def _find_mass_fix(
        self,
        ga_handle: int,
        real_id: int,
        effects: list[int],
        item_name: str,
        allow_fallback: bool,
    ) -> dict[str, Any] | None:
        strict_order = self.relic_checker.get_strictly_valid_order(real_id, effects)
        if strict_order:
            return self._mass_fix_entry(ga_handle, real_id, real_id, item_name, strict_order, False)

        valid_id = self._find_strictly_valid_relic_id(real_id, effects)
        if valid_id and valid_id != real_id:
            strict_order = self.relic_checker.get_strictly_valid_order(valid_id, effects)
            if strict_order and not self.relic_checker.check_invalidity(valid_id, strict_order):
                return self._mass_fix_entry(
                    ga_handle,
                    real_id,
                    valid_id,
                    self._relic_name(valid_id),
                    strict_order,
                    False,
                )

        if not allow_fallback:
            return None

        fallback_id = self._find_valid_relic_id_for_effects(real_id, effects)
        if fallback_id is None:
            return None

        return self._mass_fix_entry(
            ga_handle,
            real_id,
            fallback_id,
            self._relic_name(fallback_id),
            effects,
            True,
        )

    @staticmethod
    def _mass_fix_entry(
        ga_handle: int,
        old_id: int,
        new_id: int,
        new_name: str,
        effects: list[int],
        fallback: bool,
    ) -> dict[str, Any]:
        return {
            "gaHandle": ga_handle,
            "oldId": old_id,
            "newId": new_id,
            "newName": new_name,
            "effects": list(effects),
            "fallback": fallback,
        }

    def _find_valid_relic_id_for_effects(
        self,
        current_id: int,
        effects: list[int],
        target_color: int | None = None,
    ) -> int | None:
        if current_id not in self.game_data._relic_table.index:
            return None
        current_color = self.game_data._relic_table.loc[current_id, "relicColor"]
        wanted_color = current_color if target_color is None else target_color
        curses_needed = sum(
            1
            for effect_id in effects[:3]
            if effect_id not in EMPTY_EFFECT_IDS and self.game_data.effect_needs_curse(effect_id)
        )

        id_range = self.relic_checker.find_id_range(current_id)
        if not id_range:
            return None

        group_name, (range_start, range_end) = id_range
        if group_name == "illegal":
            return None

        if self.relic_checker.has_valid_order(current_id, effects):
            try:
                pools = self.game_data.relics[current_id].effect_slots
                available_curse_slots = sum(1 for pool in pools[3:] if pool != -1)
                if available_curse_slots >= curses_needed and current_color == wanted_color:
                    return current_id
            except KeyError:
                pass

        for test_id in range(range_start, range_end + 1):
            if test_id not in self.game_data._relic_table.index:
                continue
            test_color = self.game_data._relic_table.loc[test_id, "relicColor"]
            if test_color != wanted_color:
                continue
            try:
                pools = self.game_data.relics[test_id].effect_slots
                available_curse_slots = sum(1 for pool in pools[3:] if pool != -1)
                if available_curse_slots < curses_needed:
                    continue
            except KeyError:
                continue
            if self.relic_checker.has_valid_order(test_id, effects):
                return test_id

        return None

    def _find_strictly_valid_relic_id(self, current_id: int, effects: list[int]) -> int | None:
        if current_id not in self.game_data._relic_table.index:
            return None
        current_color = self.game_data._relic_table.loc[current_id, "relicColor"]
        id_range = self.relic_checker.find_id_range(current_id)
        if not id_range:
            return None

        group_name, (range_start, range_end) = id_range
        if group_name == "illegal":
            return None

        for test_id in range(range_start, range_end + 1):
            if test_id == current_id:
                continue
            if test_id not in self.game_data._relic_table.index:
                continue
            test_color = self.game_data._relic_table.loc[test_id, "relicColor"]
            if test_color != current_color:
                continue
            if self.relic_checker.get_strictly_valid_order(test_id, effects):
                return test_id

        return None

    def _relic_name(self, relic_id: int) -> str:
        relic = self.game_data.relics.get(relic_id)
        return relic.name if relic else f"Unknown ({relic_id})"

    def _validate_relic_handles(self, ga_handles: list[int]) -> list[int]:
        handles = [int(ga_handle) for ga_handle in ga_handles]
        if not handles:
            raise ValueError("No relics selected")
        missing = [ga_handle for ga_handle in handles if ga_handle not in self.inventory.relics]
        if missing:
            raise ValueError(f"Relic not found in inventory: {missing[0]}")
        return handles

    @staticmethod
    def _validate_uint32(value: int, label: str) -> None:
        if value < 0 or value > 0xFFFFFFFF:
            raise ValueError(f"{label} must be between 0 and 4294967295")

    @staticmethod
    def _validate_preset_name(name: str) -> str:
        normalized = name.strip()
        if not normalized:
            raise ValueError("Preset name is required")
        if len(normalized.encode("utf-16-le")) > 36:
            raise ValueError("Preset name must fit within 18 UTF-16 characters")
        return normalized

    @staticmethod
    def _reset_work_dir(path: Path) -> None:
        path = path.resolve()
        work_root = WORKING_DIR.resolve()
        if path.exists():
            if path == work_root or work_root not in path.parents:
                raise RuntimeError(f"Refusing to delete outside work dir: {path}")
            shutil.rmtree(path)
        path.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _name_to_path(unpack_dir: Path) -> list[tuple[str, Path]]:
        character_slots = packer.get_character_slots(unpack_dir)
        name_list: list[tuple[str, Path]] = []
        for index in range(10):
            if not character_slots[index]:
                continue
            userdata = unpack_dir / f"USERDATA_{index}"
            if not userdata.exists() or userdata.stat().st_size < 0x1000:
                continue
            name = InventoryHandler.get_player_name_from_data(userdata.read_bytes())
            if name:
                name_list.append((name, userdata.absolute()))
        return name_list
